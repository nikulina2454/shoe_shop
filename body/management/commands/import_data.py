import os
import shutil
from datetime import datetime

from django.contrib.auth.hashers import make_password
from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class Command(BaseCommand):
    help = "Импорт данных из Excel-файлов (Tovar, user_import, Заказ_import, Пункты выдачи)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path", type=str, default=".", help="Путь к папке с Excel-файлами"
        )
        parser.add_argument(
            "--images-path",
            type=str,
            default=".",
            help="Путь к папке с изображениями товаров",
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError("Установите openpyxl: pip install openpyxl")

        from body.models import (
            Category,
            Manufacturer,
            Order,
            Product,
            Role,
            Supplier,
            Unit,
            PickupPoint,
            User,
        )

        import_path = options["path"]
        images_path = options.get("images_path", import_path)

        self.stdout.write("=== Импорт данных из Excel ===")

        self._create_roles()

        dp_file = os.path.join(import_path, "Пункты выдачи_import.xlsx")
        if os.path.exists(dp_file):
            self._import_pickup_points(dp_file)

        products_file = os.path.join(import_path, "Tovar.xlsx")
        if os.path.exists(products_file):
            self._import_products(products_file, images_path)

        users_file = os.path.join(import_path, "user_import.xlsx")
        if os.path.exists(users_file):
            self._import_users(users_file)

        orders_file = os.path.join(import_path, "Заказ_import.xlsx")
        if os.path.exists(orders_file):
            self._import_orders(orders_file)

        self.stdout.write(self.style.SUCCESS("Импорт завершён успешно!"))

    def _create_roles(self):
        from body.models import Role

        roles_data = [
            (Role.CLIENT, "Авторизованный клиент"),
            (Role.MANAGER, "Менеджер"),
            (Role.ADMIN, "Администратор"),
        ]
        for role_name, _ in roles_data:
            role, created = Role.objects.get_or_create(name=role_name)
            if created:
                self.stdout.write(f"  + Роль: {role.get_name_display()}")

    def _import_pickup_points(self, filepath):
        from body.models import PickupPoint

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                address = str(row[0]).strip()
                if address:
                    PickupPoint.objects.get_or_create(address=address)
                    count += 1
        self.stdout.write(f"  Пункты выдачи: {count} записей")

    def _import_products(self, filepath, images_path):
        import shutil

        from django.conf import settings

        from body.models import Category, Manufacturer, Product, Supplier, Unit

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_data = dict(zip(headers, row))

            category, _ = Category.objects.get_or_create(
                name=str(row_data.get("Категория товара", "Без категории")).strip()
            )
            manufacturer, _ = Manufacturer.objects.get_or_create(
                name=str(row_data.get("Производитель", "Неизвестен")).strip()
            )
            supplier, _ = Supplier.objects.get_or_create(
                name=str(row_data.get("Поставщик", "Неизвестен")).strip()
            )

            discount_raw = row_data.get("Действующая скидка", 0) or 0
            try:
                discount = float(str(discount_raw).replace("%", "").strip())
            except (ValueError, TypeError):
                discount = 0.0

            try:
                price = float(str(row_data.get("Цена", 0)).replace(",", "."))
            except (ValueError, TypeError):
                price = 0.0

            try:
                stock = int(row_data.get("Кол-во на складе", 0) or 0)
            except (ValueError, TypeError):
                stock = 0

            article = str(row_data.get("Артикул", "")).strip()

            unit_name = str(row_data.get("Единица измерения", "пара") or "пара").strip()
            unit, _ = Unit.objects.get_or_create(name=unit_name)

            product, created = Product.objects.update_or_create(
                article=article,
                defaults={
                    "name": str(row_data.get("Наименование товара", "")).strip(),
                    "unit": unit,
                    "price": price,
                    "supplier": supplier,
                    "manufacturer": manufacturer,
                    "category": category,
                    "discount_percent": discount,
                    "stock_quantity": stock,
                    "description": str(
                        row_data.get("Описание товара", "") or ""
                    ).strip(),
                },
            )

            photo_name = str(row_data.get("Фото", "") or "").strip()
            if photo_name and not product.image_file:
                src_path = os.path.join(images_path, photo_name)
                if os.path.exists(src_path):
                    from django.conf import settings

                    dest_dir = os.path.join(settings.MEDIA_ROOT, "products")
                    os.makedirs(dest_dir, exist_ok=True)
                    dest_path = os.path.join(dest_dir, photo_name)
                    shutil.copy2(src_path, dest_path)
                    product.image_file = f"products/{photo_name}"
                    product.save()

            count += 1

        self.stdout.write(f"  Товары: {count} записей")

    def _import_users(self, filepath):
        from body.models import Role, User

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        role_map = {
            "Администратор": Role.ADMIN,
            "Менеджер": Role.MANAGER,
            "Авторизованный клиент": Role.CLIENT,
        }

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_data = dict(zip(headers, row))
            role_name_excel = str(row_data.get("Роль сотрудника", "")).strip()
            role_key = role_map.get(role_name_excel, Role.CLIENT)

            try:
                role = Role.objects.get(name=role_key)
            except Role.DoesNotExist:
                continue

            username = str(row_data.get("Логин", "")).strip()
            password = str(row_data.get("Пароль", "")).strip()
            full_name = str(row_data.get("ФИО", "")).strip()

            if not username:
                continue

            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    "full_name": full_name,
                    "role": role,
                    "password": make_password(password),
                },
            )
            count += 1

        self.stdout.write(f"  Пользователи: {count} записей")

    def _import_orders(self, filepath):
        from body.models import Order, OrderItem, PickupPoint, Product

        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

        status_map = {
            "Завершен": Order.STATUS_COMPLETED,
            "Новый": Order.STATUS_NEW,
            "Отменен": Order.STATUS_CANCELLED,
        }

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            row_data = dict(zip(headers, row))

            order_number = str(row_data.get("Номер заказа", "") or "").strip()
            if not order_number:
                continue

            dp_address = str(row_data.get("Адрес пункта выдачи", "") or "").strip()
            if not dp_address:
                dp_address = "Пункт выдачи (не указан)"
            pickup_point, _ = PickupPoint.objects.get_or_create(address=dp_address)

            order_date = row_data.get("Дата заказа")
            if isinstance(order_date, str):
                try:
                    order_date = datetime.strptime(order_date, "%d.%m.%Y").date()
                except ValueError:
                    self.stdout.write(
                        self.style.WARNING(
                            f'  ⚠ Заказ №{order_number}: невалидная дата заказа "{order_date}" — '
                            f"заменена на сегодняшнюю. Это ошибка в исходных данных."
                        )
                    )
                    order_date = None
            elif hasattr(order_date, "date"):
                order_date = order_date.date()

            pickup_date = row_data.get("Дата выдачи") or row_data.get("Дата доставки")
            if isinstance(pickup_date, str):
                try:
                    pickup_date = datetime.strptime(pickup_date, "%d.%m.%Y").date()
                except ValueError:
                    pickup_date = None
            elif hasattr(pickup_date, "date"):
                pickup_date = pickup_date.date()

            status_raw = str(row_data.get("Статус заказа", "Новый")).strip()
            status = status_map.get(status_raw, Order.STATUS_NEW)

            client_name = str(
                row_data.get("ФИО авторизированного клиента", "")
                or row_data.get("ФИО клиента", "")
                or ""
            ).strip()

            order, _ = Order.objects.get_or_create(
                order_number=order_number,
                defaults={
                    "order_date": order_date or datetime.now().date(),
                    "pickup_date": pickup_date,
                    "pickup_point": pickup_point,
                    "client_full_name": client_name or "Клиент (не указан)",
                    "pickup_code": str(row_data.get("Код для получения", "") or "").strip(),
                    "status": status,
                },
            )

            # Позиция заказа (в исходном шаблоне обычно одна строка = один товар).
            article = str(row_data.get("Артикул заказа", "") or row_data.get("Артикул", "") or "").strip()
            qty_raw = row_data.get("Количество", 1) or 1
            try:
                qty = int(qty_raw)
            except (ValueError, TypeError):
                qty = 1

            if article:
                product = Product.objects.filter(article=article).first()
                if product:
                    OrderItem.objects.get_or_create(order=order, product=product, defaults={"quantity": qty})
            count += 1

        self.stdout.write(f"  Заказы: {count} записей")
